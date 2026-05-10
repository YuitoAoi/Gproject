"""数据集处理服务：样本预览、图生成任务提交/查询/取消、文件下载。"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, TYPE_CHECKING

from pydantic import BaseModel, Field

from src.core.dataset import Dataset, DatasetMeta
from src.services.interfaces.dataset_repository import DatasetRepository
from src.services.interfaces.file_repository import FileRepository

if TYPE_CHECKING:
    from src.adapters.graphgen_client import GraphGenClient


# ══════════════════════════════════════════════════════════
# 样本预览
# ══════════════════════════════════════════════════════════


class SampleRequest(BaseModel):
    """样本预览请求"""

    dataset_id: int
    limit: int = Field(default=100, ge=1, le=200, description="返回行数上限")


class SampleResponse(BaseModel):
    """样本预览响应"""

    columns: List[str] = Field(default_factory=list, description="表头列名列表")
    rows: List[Dict[str, Any]] = Field(default_factory=list, description="样本数据行")
    total_rows: int = Field(default=0, description="文件总行数")
    error: Optional[str] = None


# ══════════════════════════════════════════════════════════
# 图生成任务
# ══════════════════════════════════════════════════════════


class DatasetProcessRequest(BaseModel):
    dataset_id: int
    """图生成任务请求。input_file 由 service 层根据 dataset_id 自动填入。"""
    model_config = {"extra": "forbid"}

    # LLM
    api_key: str
    synthesizer_url: str
    synthesizer_model: str
    mode: Literal[
        "atomic",
        "multi_hop",
        "aggregated",
        "CoT",
        "multi_choice",
        "multi_answer",
        "fill_in_blank",
        "true_false",
    ]
    data_format: Literal["Alpaca", "Sharegpt", "ChatML"]

    # Content field mapping — which column to use as source text
    content_field: str = "content"

    # Optional with defaults
    tokenizer: str = "cl100k_base"
    trainee_model: Optional[str] = None
    trainee_url: Optional[str] = None
    trainee_api_key: Optional[str] = None
    chunk_size: int = Field(default=1024, gt=0)
    chunk_overlap: int = Field(default=100, ge=0)
    quiz_samples: int = Field(default=2, ge=0)
    partition_method: Literal["dfs", "bfs", "leiden", "ece"] = "ece"
    dfs_max_units: int = 5
    bfs_max_units: int = 5
    leiden_max_size: int = 20
    leiden_use_lcc: bool = False
    leiden_random_seed: int = 42
    ece_max_units: int = 20
    ece_min_units: int = 3
    ece_max_tokens: int = 10240
    ece_unit_sampling: Literal["random", "max_loss", "min_loss"] = "random"
    rpm: int = Field(default=1000, gt=0)
    tpm: int = Field(default=50000, gt=0)


class DatasetProcessResponse(BaseModel):
    """图生成任务状态响应"""

    job_id: str
    status: Literal["pending", "running", "done", "failed", "cancelled"]
    created_at: Optional[str] = None
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    progress: float = Field(default=0.0, ge=0.0, le=1.0)
    error: Optional[str] = None
    output_path: Optional[str] = None


# ══════════════════════════════════════════════════════════
# 服务
# ══════════════════════════════════════════════════════════


class DatasetProcessService:
    """数据集处理服务。

    依赖通过构造函数注入。
    """

    def __init__(
        self,
        dataset_repo: DatasetRepository,
        file_repo: FileRepository,
        gg_client: GraphGenClient,
        celery_client=None,
        dataset_log_repo=None,
        task_repo=None,
    ) -> None:
        self._dataset_repo = dataset_repo
        self._file_repo = file_repo
        self._gg = gg_client
        self._celery = celery_client
        self._dataset_log_repo = dataset_log_repo
        self._task_repo = task_repo

    # ── 样本预览 ──────────────────────────────────────────────

    def get_sample(self, request: SampleRequest) -> SampleResponse:
        """读取数据集文件，返回前 N 条样本及表头。"""
        ds = self._dataset_repo.find_by_id(request.dataset_id)
        if ds is None:
            return SampleResponse(error=f"Dataset not found: {request.dataset_id}")

        file_path = ds.meta.file_path
        if not self._file_repo.exists(file_path):
            return SampleResponse(error=f"File not found: {file_path}")

        fmt = ds.meta.format
        try:
            if fmt == "csv":
                return self._sample_csv(file_path, request.limit)
            elif fmt in ("json", "jsonl"):
                if ds.status == 0:
                    return self._sample_raw_text(file_path, request.limit)
                result = self._sample_jsonl(file_path, request.limit)
                if not result.rows and not result.error:
                    result = self._sample_json_array(file_path, request.limit)
                if not result.rows and not result.error:
                    result = self._sample_raw_text(file_path, request.limit)
                return result
            elif fmt == "xlsx":
                return self._sample_xlsx(file_path, request.limit)
            elif fmt in ("txt", "md"):
                return self._sample_raw_text(file_path, request.limit)
            else:
                return SampleResponse(error=f"Unsupported format: {fmt}")
        except Exception as exc:
            return SampleResponse(error=f"Sample read error: {exc}")

    # ── 图生成任务 ────────────────────────────────────────────

    def process(self, request: DatasetProcessRequest) -> DatasetProcessResponse:
        """提交图生成任务到 GraphGen。仅 status=0（未处理）的数据集可提交。"""
        ds = self._dataset_repo.find_by_id(request.dataset_id)
        if ds is None:
            return DatasetProcessResponse(
                job_id="",
                status="failed",
                error=f"Dataset not found: {request.dataset_id}",
            )

        if ds.status != 0:
            return DatasetProcessResponse(
                job_id="",
                status="failed",
                error=f"Dataset already processed (status={ds.status})",
            )

        file_path = ds.meta.file_path
        if not self._file_repo.exists(file_path):
            return DatasetProcessResponse(
                job_id="",
                status="failed",
                error=f"File not found: {file_path}",
            )

        input_file = Path(file_path).name
        payload = request.model_dump(exclude={"dataset_id"})
        payload["input_file"] = input_file

        resp = self._gg.submit_job(payload)
        if resp.is_error:
            return DatasetProcessResponse(
                job_id="",
                status="failed",
                error=self._parse_error(resp),
            )

        data = resp.json()
        job_id = data.get("job_id", "")

        # 创建任务记录
        if job_id and self._task_repo is not None:
            from src.core.task_record import TaskRecord as TR
            task = TR(
                owner_id=ds.owner_id,
                task_name=ds.name,
                task_type="cleaning",
                status="pending",
                config=TR.config_to_json({
                    "job_id": job_id,
                    "dataset_id": request.dataset_id,
                    "mode": request.mode,
                    "data_format": request.data_format,
                    "content_field": request.content_field,
                }),
            )
            self._task_repo.insert(task)

        # 提交 Celery 异步监控（Celery 不可用时不影响任务提交）
        if job_id and self._celery is not None:
            try:
                self._celery.send_task(
                    "dataset.monitor_graphgen",
                    kwargs={"job_id": job_id, "dataset_id": request.dataset_id},
                )
            except Exception:
                pass

        return DatasetProcessResponse(
            job_id=job_id,
            status=data.get("status", "pending"),
        )

    def check_job(self, job_id: str, dataset_id: int) -> DatasetProcessResponse:
        """查询 GraphGen 任务状态，完成后自动更新数据集 status 和 output_path。

        status 流转: 0(未处理) → 1(处理中/已提交) → 2(已完成) / -1(失败)
        """
        result = self.get_job(job_id)
        ds = self._dataset_repo.find_by_id(dataset_id)
        if ds is None:
            return result

        if result.status in ("pending", "running") and ds.status == 0:
            ds.status = 1
            self._dataset_repo.update(dataset_id, ds)
        elif result.status == "done":
            ds.status = 0
            if result.output_path:
                ds.meta.output_path = result.output_path
            self._dataset_repo.update(dataset_id, ds)

            output_path = result.output_path
            if output_path and os.path.isfile(output_path):
                stat = os.stat(output_path)
                ext = os.path.splitext(output_path)[1].lstrip(".").lower()
                if ext not in ("txt", "md", "csv", "xlsx", "json", "jsonl"):
                    ext = "jsonl"

                pattern = re.compile(rf"^{re.escape(ds.name)}_清洗(\d+)$")
                max_seq = 0
                for d in self._dataset_repo.find_by_owner(ds.owner_id):
                    m = pattern.match(d.name)
                    if m:
                        seq = int(m.group(1))
                        if seq > max_seq:
                            max_seq = seq

                next_seq = max_seq + 1
                output_ds = Dataset.new(
                    owner_id=ds.owner_id,
                    name=f"{ds.name}_清洗{next_seq}",
                    desc=f"GraphGen 清洗产物（第{next_seq}次）",
                    meta=DatasetMeta(
                        format=ext,
                        file_path=output_path,
                        file_size=stat.st_size,
                    ),
                    status=2,
                    tag_ids=list(ds.tag_ids),
                )
                output_ds.created_at = datetime.fromtimestamp(stat.st_ctime)
                output_ds.updated_at = datetime.fromtimestamp(stat.st_ctime)
                self._dataset_repo.create(output_ds)

                if self._dataset_log_repo is not None:
                    log_src = os.path.join("cache", "logs", f"{job_id}.log")
                    log_dst_dir = os.path.join("dataset", "logs", "dataset_logs")
                    os.makedirs(log_dst_dir, exist_ok=True)
                    log_dst = os.path.join(log_dst_dir, f"{job_id}.log")
                    if os.path.isfile(log_src):
                        shutil.copy2(log_src, log_dst)
                        output_ds.meta.log_path = log_dst
                        self._dataset_repo.update(output_ds.id, output_ds)
                    from src.core.dataset_log import DatasetLog as DSLog
                    log_record = DSLog(
                        job_id=job_id,
                        dataset_id=output_ds.id,
                        log_path=output_ds.meta.log_path or "",
                        created_at=datetime.now(),
                    )
                    self._dataset_log_repo.insert(log_record)

                if self._task_repo is not None:
                    task = self._task_repo.find_by_config_job_id(ds.owner_id, job_id)
                    if task:
                        from src.core.task_record import TaskRecord as TR
                        task.status = "done"
                        task.progress = 1.0
                        task.phase = "处理完成"
                        task.config = TR.config_to_json({
                            "job_id": job_id,
                            "dataset_id": dataset_id,
                            "output_dataset_id": output_ds.id,
                            "mode": result.status,
                        })
                        task.updated_at = datetime.now()
                        self._task_repo.update(task.id, task)

        elif result.status == "failed":
            ds.status = 0
            self._dataset_repo.update(dataset_id, ds)
            if self._task_repo is not None:
                task = self._task_repo.find_by_config_job_id(ds.owner_id, job_id)
                if task:
                    task.status = "failed"
                    task.phase = result.error or "任务失败"
                    task.updated_at = datetime.now()
                    self._task_repo.update(task.id, task)

        return result

    def get_job(self, job_id: str) -> DatasetProcessResponse:
        """查询 GraphGen 任务状态。"""
        resp = self._gg.get_job(job_id)
        if resp.is_error:
            return DatasetProcessResponse(
                job_id=job_id,
                status="failed",
                error=self._parse_error(resp),
            )

        data = resp.json()
        return DatasetProcessResponse(
            job_id=data.get("job_id", job_id),
            status=data.get("status", "unknown"),
            created_at=data.get("created_at"),
            started_at=data.get("started_at"),
            finished_at=data.get("finished_at"),
            progress=data.get("progress", 0.0),
            error=data.get("error"),
            output_path=data.get("output_path"),
        )

    def cancel_job(self, job_id: str) -> DatasetProcessResponse:
        """取消 GraphGen 任务。"""
        resp = self._gg.cancel_job(job_id)
        if resp.is_error:
            return DatasetProcessResponse(
                job_id=job_id,
                status="failed",
                error=self._parse_error(resp),
            )

        data = resp.json()
        return DatasetProcessResponse(
            job_id=data.get("job_id", job_id),
            status=data.get("status", "cancelled"),
        )

    # ── 内部 ──────────────────────────────────────────────────

    @staticmethod
    def _parse_error(resp) -> str:
        """从 httpx Response 提取错误信息。"""
        try:
            detail = resp.json()
            if isinstance(detail, dict):
                return detail.get("detail", resp.text)
            if isinstance(detail, list):
                return "; ".join(
                    d.get("msg", str(d)) for d in detail if isinstance(d, dict)
                )
            return str(detail)
        except Exception:
            return resp.text or f"HTTP {resp.status_code}"

    # ── 样本读取 ──────────────────────────────────────────────

    def _sample_csv(self, file_path: str, limit: int) -> SampleResponse:
        raw = self._file_repo.read(file_path)
        reader = csv.reader(io.StringIO(raw.decode("utf-8")))
        try:
            columns = next(reader)
        except StopIteration:
            return SampleResponse(columns=[], rows=[], total_rows=0)

        rows: List[Dict[str, Any]] = []
        total = 0
        for row in reader:
            total += 1
            if len(rows) < limit:
                rows.append(
                    {
                        col: row[i] if i < len(row) else ""
                        for i, col in enumerate(columns)
                    }
                )
        return SampleResponse(
            columns=columns,
            rows=rows,
            total_rows=total + 1,  # +1 for header
        )

    def _sample_jsonl(self, file_path: str, limit: int) -> SampleResponse:
        raw = self._file_repo.read(file_path)
        lines = raw.decode("utf-8").strip().splitlines()
        rows: List[Dict[str, Any]] = []
        columns: List[str] = []
        for line in lines:
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    rows.append(obj)
                    for k in obj:
                        if k not in columns:
                            columns.append(k)
                    if len(rows) >= limit:
                        break
            except json.JSONDecodeError:
                continue
        return SampleResponse(
            columns=columns,
            rows=rows[:limit],
            total_rows=len(lines),
        )

    def _sample_json_array(self, file_path: str, limit: int) -> SampleResponse:
        """解析标准 JSON 数组 [{...}, {...}] 或单对象 {...}。"""
        raw = self._file_repo.read(file_path)
        try:
            data = json.loads(raw.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return SampleResponse(columns=[], rows=[], total_rows=0)

        if isinstance(data, dict):
            return SampleResponse(
                columns=list(data.keys()),
                rows=[data],
                total_rows=1,
            )

        if not isinstance(data, list):
            return SampleResponse(columns=[], rows=[], total_rows=0)

        rows: List[Dict[str, Any]] = []
        columns: List[str] = []
        for item in data[:limit]:
            if isinstance(item, dict):
                rows.append(item)
                for k in item:
                    if k not in columns:
                        columns.append(k)
        return SampleResponse(
            columns=columns,
            rows=rows,
            total_rows=len(data),
        )

    def _sample_xlsx(self, file_path: str, limit: int) -> SampleResponse:
        raw = self._file_repo.read(file_path)
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return SampleResponse(columns=[], rows=[], total_rows=0)
        rows_iter = ws.iter_rows(values_only=True)

        try:
            columns = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            wb.close()
            return SampleResponse(columns=[], rows=[], total_rows=0)

        rows: List[Dict[str, Any]] = []
        total = 0
        for row in rows_iter:
            total += 1
            if len(rows) < limit:
                rows.append(
                    {
                        col: (row[i] if i < len(row) else "")
                        for i, col in enumerate(columns)
                    }
                )
        wb.close()
        return SampleResponse(
            columns=columns,
            rows=rows,
            total_rows=total + 1,
        )

    def _sample_raw_text(self, file_path: str, limit: int) -> SampleResponse:
        raw = self._file_repo.read(file_path)
        text = raw.decode("utf-8", errors="replace")
        lines = [l for l in text.splitlines() if l.strip()]
        columns = ["content"]
        rows = [{"content": line} for _, line in zip(range(limit), lines)]
        return SampleResponse(columns=columns, rows=rows, total_rows=len(lines))
