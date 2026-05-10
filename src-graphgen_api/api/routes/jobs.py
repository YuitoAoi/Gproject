import json
import os
from datetime import datetime, timezone

from fastapi import APIRouter, BackgroundTasks, HTTPException

from api.config import config
from api.job_manager import JobManager
from api.schemas.generation import GenerateRequest, GenerateResponse, JobStatus
from api.services.pipeline import execute_pipeline

router = APIRouter(prefix="/api/v1", tags=["jobs"])


_manager: JobManager | None = None


def _get_manager() -> JobManager:
    global _manager
    if _manager is None:
        _manager = JobManager(config.JOBS_DIR)
    return _manager


def _resolve_input(input_file: str) -> str:
    """Resolve input_file relative to datasets_dir. Rejects path traversal."""
    datasets = os.path.abspath(config.DATASETS_DIR)
    os.makedirs(datasets, exist_ok=True)
    resolved = os.path.normpath(os.path.join(datasets, input_file))
    if not resolved.startswith(datasets):
        raise HTTPException(
            status_code=422,
            detail=f"Invalid input_file (path traversal denied): {input_file}",
        )
    return resolved


def _validate_input_file(path: str) -> str | None:
    """Return error message if file is invalid, None if OK."""
    if not os.path.isfile(path):
        return f"Input file not found: {path}"
    if os.path.getsize(path) == 0:
        return f"Input file is empty: {path}"

    ext = os.path.splitext(path)[1].lstrip(".").lower()
    if ext not in ("json", "jsonl", "txt", "csv", "md", "xlsx"):
        return f"Unsupported file type: {ext}"

    if ext in ("json", "jsonl"):
        return _validate_json_format(path)
    if ext == "txt":
        with open(path, "r", encoding="utf-8") as f:
            head = f.read(100)
        if not head.strip():
            return f"Input file has no readable content: {path}"
    if ext == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            if not wb.sheetnames:
                return f"Input xlsx has no sheets: {path}"
            ws = wb.active
            if ws is None:
                return f"Input xlsx has no active sheet: {path}"
            if ws.max_row < 1 or ws.max_column < 1:
                return f"Input xlsx has no readable content: {path}"
            wb.close()
        except Exception as e:
            return f"Input xlsx is invalid: {e}"

    return None


def _validate_json_format(path: str) -> str | None:
    """Validate JSON/JSONL file has valid JSON lines and at least some data."""
    with open(path, "r", encoding="utf-8") as f:
        head = f.read(200).strip()
    if not head:
        return f"Input file has no content (empty first line): {path}"

    if head.startswith("["):
        try:
            data = json.loads(head)
        except json.JSONDecodeError:
            try:
                with open(path, "r", encoding="utf-8") as f2:
                    data = json.load(f2)
            except json.JSONDecodeError as e:
                return f"Input file has invalid JSON array: {e}"
        if not isinstance(data, list) or len(data) == 0:
            return f"Input JSON array is empty or not a list"
        first_row = data[0]
    else:
        first_line = head.split("\n")[0]
        try:
            first_row = json.loads(first_line)
        except json.JSONDecodeError as e:
            return f"Input file has invalid JSON on line 1: {e}"

    if not isinstance(first_row, dict):
        return f"Input file must contain JSON objects, got: {type(first_row).__name__}"

    return None


@router.post("/jobs", response_model=GenerateResponse, status_code=202)
async def create_job(req: GenerateRequest, background_tasks: BackgroundTasks):
    input_path = _resolve_input(req.input_file)
    error = _validate_input_file(input_path)
    if error:
        raise HTTPException(status_code=422, detail=error)

    manager = _get_manager()
    job_id = manager.create(req.model_dump())
    background_tasks.add_task(execute_pipeline, job_id, req, input_path)
    return GenerateResponse(job_id=job_id)


@router.get("/jobs/{job_id}", response_model=JobStatus)
async def get_job(job_id: str):
    manager = _get_manager()
    try:
        data = manager.get(job_id)
        return JobStatus(**data)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")


@router.delete("/jobs/{job_id}")
async def cancel_job(job_id: str):
    manager = _get_manager()
    now = datetime.now(timezone.utc).isoformat()

    # Try to cancel from pending state
    if manager.update_if(
        job_id,
        condition={"status": "pending"},
        updates={"status": "cancelled", "finished_at": now, "error": "Cancelled by user"},
    ):
        return {"job_id": job_id, "status": "cancelled"}

    # Try to cancel from running state
    if manager.update_if(
        job_id,
        condition={"status": "running"},
        updates={"status": "cancelled", "finished_at": now, "error": "Cancelled by user"},
    ):
        return {"job_id": job_id, "status": "cancelled"}

    # Job is already done/failed/cancelled, or doesn't exist
    try:
        data = manager.get(job_id)
        return {"job_id": job_id, "status": data["status"]}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")


@router.get("/jobs/{job_id}/logs")
async def get_job_logs(job_id: str, offset: int = 0):
    """返回 job 日志文件中 offset 之后的新行。

    每次返回文件从 offset 到末尾的新内容，同时返回新的文件末尾位置。
    前端/Celery 可用返回的 offset 做增量轮询。
    """
    log_file = os.path.join(config.LOG_DIR, f"{job_id}.log")
    if not os.path.isfile(log_file):
        return {"lines": [], "offset": 0}

    with open(log_file, "r", encoding="utf-8") as f:
        f.seek(0, os.SEEK_END)
        end_pos = f.tell()
        if offset < 0 or offset > end_pos:
            offset = max(0, end_pos - 8192)
        f.seek(offset)
        if offset > 0:
            f.readline()
        lines = [line.rstrip("\n") for line in f.readlines()]

    return {"lines": lines, "offset": end_pos}
